import csv
import sqlite3


def load_from_csv(
    filepath,
    name_column="Contacted Person Name",
    email_column="Institution Email",
    status_column="Status",
    institution_column="Institution Name",
):
    """
    Load recipients from the AIfSI KPI Dashboard CSV.

    Expected columns include configurable name/email fields and 'Status'.
    Only rows whose Status column is **blank** (empty) are included.
    Rows with any status text at all are skipped.
    Stops collecting as soon as a row has no name (end of useful data).
    """
    recipients = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get(name_column, "").strip()
            email = row.get(email_column, "").strip()
            status = row.get(status_column, "").strip()
            institution = row.get(institution_column, "").strip()

            # If the name field is empty we've hit the end of real data
            if not name:
                break

            # Only include rows where Status is completely blank
            if status != "":
                continue

            if not email:
                continue  # has a name but no email — skip

            recipients.append({
                "full_name": name,
                "email": email,
                "institution_name": institution,
            })
    return recipients


def load_from_sqlite(db_path, query="SELECT full_name, email FROM contacts"):
    """Load recipients from a SQLite database."""
    recipients = []
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        for row in conn.execute(query):
            recipients.append({
                "full_name": row["full_name"],
                "email": row["email"],
            })
    return recipients


def load_from_excel(filepath):
    """Load recipients from an Excel file with columns: full_name, email"""
    import openpyxl

    recipients = []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    name_idx = headers.index("full_name")
    email_idx = headers.index("email")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[email_idx]:  # skip empty rows
            recipients.append({
                "full_name": row[name_idx],
                "email": row[email_idx],
            })
    wb.close()
    return recipients